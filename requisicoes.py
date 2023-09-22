from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from time import sleep
from bs4 import BeautifulSoup
import pyautogui
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

options = Options();

options.add_argument('window-size=1280,768')

servico = Service(ChromeDriverManager().install())
navegador = webdriver.Chrome(service=servico, options=options)

navegador.get("https://app.melibox.com.br/advProductPosition")

inputEmail = navegador.find_element(By.NAME, "email");
inputSenha = navegador.find_element(By.NAME, "password");

inputEmail.send_keys("Nelson.wilhelms@gmail.com");
inputSenha.send_keys("caraio123");
inputSenha.submit();

sleep(3)

inputTagProduto = navegador.find_element(By.NAME, "textoPesquisa");
inputTagProduto.send_keys("Brinquedo Importado")
buttonEnviar = navegador.find_element(By.XPATH, "//*[@id='content']/div/div/div/div/div[1]/div/div/button[1]")
buttonEnviar.click();


sleep(4)


idAnuncios = []
data = []
diasAtivo = []
visitasTotais = []
visitasDiarias = []
taxaConversao = []
titulos = []
valor = []
url = []



pyautogui.scroll(-10000, x=100, y=100)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)
pyautogui.scroll(-10000)
sleep(0.5)



pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)
pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)
pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)
sleep(0.5)
pyautogui.scroll(+10000)



i = 2;
while(i < 50):
    print("teste", i)
    sleep(0.7);
    xpath = f'//*[@id="analiticoFindPosAdViewRetorno"]/table/tbody/tr[{i}]/td[5]/small'
    vendas = navegador.find_element(By.XPATH, xpath).text
    if vendas == 'De 26 a 50' or vendas == 'De 51 a 100' or vendas == 'De 5 a 25' or vendas == 'De 101 a 150' or vendas == 'De 151 a 250' or vendas == 'De 251 a 500' or vendas == 'De 501 a 5000':
        print("encontrou")
        xpathImg = f'//*[@id="analiticoFindPosAdViewRetorno"]/table/tbody/tr[{i}]/td[9]/small[2]/img'
        xpathUrl = f'//*[@id="analiticoFindPosAdViewRetorno"]/table/tbody/tr[{i}]/td[9]/small[4]/a'
        xpathTitulo = f'//*[@id="analiticoFindPosAdViewRetorno"]/table/tbody/tr[{i}]/td[2]/small'
        xpathId = f'//*[@id="analiticoFindPosAdViewRetorno"]/table/tbody/tr[{i}]/td[1]/small'
        
        titulo = navegador.find_element(By.XPATH, xpathTitulo)
        idAnuncio = navegador.find_element(By.XPATH, xpathId)
        print("teste ", idAnuncio.text)

        elementoUrl = navegador.find_element(By.XPATH, xpathUrl).get_attribute('href')
        elemento_img = navegador.find_element(By.XPATH, xpathImg)
        elemento_img.click()
        print("clicou")
        sleep(5)
        sleep(0.7)
        sleep(0.7)
        sleep(0.7)
        sleep(0.7)
        sleep(0.7)
        elemento_info = navegador.find_element(By.XPATH, '//*[@id="btnResumo"]');
        elemento_info.click()
        sleep(2)


        page_content = navegador.page_source
        site = BeautifulSoup(page_content, 'html.parser')

       

        paragrafoDescricao = site.find('p' ,attrs={'style' :'font-size: 13px;'})

        infosImportantes = paragrafoDescricao.find_all('strong')
        print(infosImportantes)

        if len(infosImportantes) < 4:
            pyautogui.press('esc');
            
        else:
            porcentagem = infosImportantes[4].find('span', attrs={"color" :"green"});

            if porcentagem == None:
                    porcentagem = infosImportantes[4].find('span', attrs={"color" :"red"});

            idAnuncios.append(idAnuncio.text)
            titulos.append(titulo.text);
            valor.append(site.find('span', attrs={'style': 'color:white;'}))
            data.append(infosImportantes[0])
            diasAtivo.append(infosImportantes[1])
            visitasTotais.append(infosImportantes[2])
            visitasDiarias.append(infosImportantes[3])
            taxaConversao.append(porcentagem.decode_contents())
            url.append(elementoUrl)

            pyautogui.press('esc')
            
          # scroll down 10 "clicks"

        i = i + 1;
    else:
        print("nao encontro")
        i = i+ 1;
print("fim")

j = 1
m = 3
i = 1
while(j < 20):
    k = m;

    if idAnuncio.text == '949':
        break;

    while(k < 52):
        sleep(0.5)
        sleep(0.5)
        sleep(0.5)
        xpath = f'//*[@id="lazyloading"]/table[{j}]/tbody/tr[{k}]/td[5]/small'
        vendas = navegador.find_element(By.XPATH, xpath).text

        xpathId = f'//*[@id="lazyloading"]/table[{j}]/tbody/tr[{k}]/td[1]/small'
        idAnuncio = navegador.find_element(By.XPATH, xpathId);

        print("id: ", idAnuncio.text);

        if idAnuncio.text == '949':
             break;

        if vendas == 'De 26 a 50' or vendas == 'De 51 a 100' or vendas == 'De 5 a 25' or vendas == 'De 101 a 150' or vendas == 'De 151 a 250' or vendas == 'De 251 a 500' or vendas == 'De 501 a 5000':
            print("encontrou")
            print(vendas)
            xpathImg = f'//*[@id="lazyloading"]/table[{j}]/tbody/tr[{k}]/td[9]/small[2]/img'
            xpathUrl = f'//*[@id="lazyloading"]/table[{j}]/tbody/tr[{k}]/td[9]/small[4]/a'
            xpathTitulo = f'//*[@id="lazyloading"]/table[{j}]/tbody/tr[{k}]/td[2]/small'
            xpathId = f'//*[@id="lazyloading"]/table[{j}]/tbody/tr[{k}]/td[1]/small'
        
            titulo = navegador.find_element(By.XPATH, xpathTitulo) 
            idAnuncio = navegador.find_element(By.XPATH, xpathId)
            print("teste ", idAnuncio.text)
            
            elementoUrl = navegador.find_element(By.XPATH, xpathUrl).get_attribute('href')
            elemento_img = navegador.find_element(By.XPATH, xpathImg)
            elemento_img.click()
            sleep(5)
            sleep(0.7)
            sleep(0.7)
            sleep(0.7)
            sleep(0.7)
            sleep(0.7)
            sleep(0.7)
            sleep(0.7)
            sleep(0.7)
            sleep(0.7)
            sleep(0.7)
            sleep(0.7)
            
            idNumero = int(idAnuncio.text);
            if(idNumero > 200):
                 sleep(0.7);
            if(idNumero > 400):
                 sleep(0.7);
            if(idNumero > 600):
                 sleep(0.7);
            if(idNumero > 800):
                 sleep(0.7);
            

            elemento_info = navegador.find_element(By.XPATH, '//*[@id="btnResumo"]');
            elemento_info.click()
            sleep(2)
            
            page_content = navegador.page_source
            site = BeautifulSoup(page_content, 'html.parser')


            paragrafoDescricao = site.find('p' ,attrs={'style' :'font-size: 13px;'})

            infosImportantes = paragrafoDescricao.find_all('strong')

            if len(infosImportantes) < 4:
                pyautogui.press('esc');
            else:
                porcentagem = infosImportantes[4].find('span', attrs={"color" :"green"});

                if porcentagem == None:
                        porcentagem = infosImportantes[4].find('span', attrs={"color" :"red"});

                idAnuncios.append(idAnuncio.text)
                titulos.append(titulo.text);
                valor.append(site.find('span', attrs={'style': 'color:white;'}))
                data.append(infosImportantes[0])
                diasAtivo.append(infosImportantes[1])
                visitasTotais.append(infosImportantes[2])
                visitasDiarias.append(infosImportantes[3])
                taxaConversao.append(porcentagem.decode_contents())
                url.append(elementoUrl)

                pyautogui.press('esc')

            
            print("----------")
            print(elementoUrl)
            print("----------")


            i = i + 1;
            k = k + 1
        else:
            print("nao encontro")
            i = i+ 1;
            k = k + 1;
    j = j + 1
    m = m + 1


print("fim")


sleep(3)



i = 0;


workbook = openpyxl.Workbook()
sheet = workbook.active

headers = ["ID", "Título", "Preço", "Data", "Dias Ativos", "Total de Visitas", "Visitas Diárias", "Taxa de Conversão", "URL"]

for col_num, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col_num)
    cell.value = header

# Insira os dados nas células da planilha
for row_num, (id_anuncio, titulo, preco, data, dias_ativos, total_visitas, visitas_diarias, taxa_conversao, url) in enumerate(
    zip(idAnuncios, titulos, [valor[i].decode_contents() for i in range(len(valor))], [data[i].decode_contents() for i in range(len(data))],
        [diasAtivo[i].decode_contents() for i in range(len(diasAtivo))], [visitasTotais[i].decode_contents() for i in range(len(visitasTotais))],
        [visitasDiarias[i].decode_contents() for i in range(len(visitasDiarias))], [taxaConversao[i] for i in range(len(taxaConversao))], url), start=1):

    data_row = [id_anuncio, titulo, preco, data, dias_ativos, total_visitas, visitas_diarias, taxa_conversao, url]

    for col_num, value in enumerate(data_row, 1):
        cell = sheet.cell(row=row_num + 1, column=col_num)
        cell.value = value

# Salve o arquivo Excel
workbook.save("Brinquedo Importado.xlsx")

# Feche o arquivo Excel
workbook.close()

input("aa")